import sys

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils import term_str, dept_codes_map, vault_syllabi_hierarchy_url

wb = load_workbook(sys.argv[1])
# only one worksheet
ws: Worksheet = wb.worksheets[0]
# skip header and "Academic Affiliates" 2nd row
semester = term_str()
for row in ws.iter_rows(min_row=4, values_only=True):
    # coerce to string just to silence type warnings. these will be all strings
    dept, names, emails = [str(cell) for cell in row[:3]]
    names = names.replace("\n\n", " & ")
    emails = emails.replace("\n\n", ", ")
    # skip division lines
    if "Division" in dept:
        continue
    # TODO have dept codes and parent division in the map
    depts = dept_codes_map[dept]
    links = "\n".join(
        vault_syllabi_hierarchy_url("Fine Arts Division", dept_code, semester)
        for dept_code in depts
    )
    print(
        f"""
To: {emails}

Hello {names},

Here are the current {dept} syllabi for {semester}:
{links}
"""
    )
